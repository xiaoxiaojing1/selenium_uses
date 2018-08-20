
from openpyxl import load_workbook

class ReadOpenData:
    """
    读取excel表结构
    """
    def __init__(self):
        "Constructor of the object"
        self.workbook = load_workbook(r"D:\selenium_uses\test_data\data.xlsx")

    def get_sheet(self):
        "读取sheet部分"
        return self.workbook.sheetnames[0]

    def get_cell_dire(self,row,col):
        """
        通过坐标直接读取单元格
        :param row: int
        :param col: int
        :return: 对应row & col 的 cell
        """
        row+=1
        col+=1
        return self.workbook.active.cell(row, col).value

    def get_rows(self,row):
        """
        读取某一行数据
        :param row: int
        :return:包含这一行数据的list  class:openpyxl.cell
        """
        values = []
        cells = list(self.workbook.active.iter_rows())[row]
        for cell in cells:
            values.append(cell.value)
        return values #[v1,v2,v3]


    def get_cell_by_row(self,row,col):
        """
        按行读取单元格
        :param row: int
        :param col: int
        :return: str, 对应row & col 数据  class:openpyxl.cell
        """
        return self.get_rows(row)[col]
